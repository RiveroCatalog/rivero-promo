import zipfile, xml.etree.ElementTree as ET, os, re, json, base64, subprocess, tempfile
from pathlib import Path
from openpyxl import load_workbook

import os
TOKEN = os.environ.get('GH_TOKEN', '')
REPO  = 'RiveroCatalog/rivero-promo'
API   = f'https://api.github.com/repos/{REPO}/contents'

def curl_upload(path, data):
    tmp = tempfile.mktemp(suffix='.json')
    payload = {'message': f'Add {path}', 'content': base64.b64encode(data).decode()}
    with open(tmp, 'w') as f:
        json.dump(payload, f)
    r = subprocess.run(['curl','-s','-X','PUT',
        '-H', f'Authorization: token {TOKEN}',
        '-H', 'Content-Type: application/json',
        f'{API}/{path}', '-d', f'@{tmp}'],
        capture_output=True, text=True)
    os.unlink(tmp)
    d = json.loads(r.stdout)
    return 'commit' in d

r = subprocess.run(['curl','-s','-H',f'Authorization: token {TOKEN}',
    f'https://api.github.com/repos/{REPO}/git/trees/main?recursive=1'],
    capture_output=True, text=True)
tree = json.loads(r.stdout).get('tree', [])
existing = set(
    re.sub(r'\.(png|jpg|jpeg)$','', item['path'].replace('products/',''))
    for item in tree if item['path'].startswith('products/')
)
print(f'Imagenes ya en repo: {len(existing)}')

xlsx = sorted(Path.home().glob('Downloads/*.xlsx'), key=lambda x: x.stat().st_size, reverse=True)[0]
print(f'Excel: {xlsx.name}')

wb = load_workbook(xlsx, read_only=True, data_only=True)
ws = wb.active
row_to_sku = {}
for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
    sku = str(row[0]).strip() if row[0] else ''
    if sku and sku != 'None':
        row_to_sku[i + 1] = sku
print(f'SKUs en Excel: {len(row_to_sku)}')

uploaded = skipped = failed = 0

with zipfile.ZipFile(xlsx) as z:
    names = z.namelist()
    drawings = [n for n in names if re.match(r'xl/drawings/drawing\d+\.xml$', n)]
    print(f'Drawings encontrados: {len(drawings)}')
    for drawing_path in drawings:
        rels_path = drawing_path.replace('drawings/', 'drawings/_rels/') + '.rels'
        if rels_path not in names:
            continue
        rels = ET.fromstring(z.read(rels_path))
        rid_to_img = {rel.get('Id'): os.path.basename(rel.get('Target','')) for rel in rels}
        draw = ET.fromstring(z.read(drawing_path))
        ns_xdr = 'http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing'
        ns_a   = 'http://schemas.openxmlformats.org/drawingml/2006/main'
        ns_r   = 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'
        for anchor in draw.findall(f'.//{{{ns_xdr}}}oneCellAnchor') + draw.findall(f'.//{{{ns_xdr}}}twoCellAnchor'):
            from_el = anchor.find(f'{{{ns_xdr}}}from')
            pic     = anchor.find(f'.//{{{ns_xdr}}}pic')
            if from_el is None or pic is None: continue
            row_el = from_el.find(f'{{{ns_xdr}}}row')
            blip   = pic.find(f'.//{{{ns_a}}}blip')
            if row_el is None or blip is None: continue
            draw_row = int(row_el.text)
            rid = blip.get(f'{{{ns_r}}}embed')
            img_name = rid_to_img.get(rid, '')
            img_path = f'xl/media/{img_name}'
            if img_path not in names: continue
            sku = row_to_sku.get(draw_row) or row_to_sku.get(draw_row - 1)
            if not sku: continue
            safe_sku = re.sub(r'[^A-Za-z0-9._-]', '_', sku)
            if safe_sku in existing:
                skipped += 1
                continue
            ext = os.path.splitext(img_name)[1] or '.png'
            img_data = z.read(img_path)
            if curl_upload(f'products/{safe_sku}{ext}', img_data):
                uploaded += 1
                existing.add(safe_sku)
                if uploaded % 10 == 0:
                    print(f'  {uploaded} subidas...')
            else:
                failed += 1

print(f'\nListo: {uploaded} subidas | {skipped} ya existian | {failed} fallaron')
